import os
import openpyxl


class RemainsService:
    _tmpl_excel_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'template.xlsx')
    _res_excel_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'result.xlsx')
    _data = {}

    # def __init__(self):
    #     if os.path.exists(self._res_excel_path):
    #         os.remove(self._res_excel_path)

    def clean(self):
        self._data = {}

    def set(self, barcode, value):
        self._data[barcode] = value

    def get(self, barcode):
        self._data.get(barcode)

    def save(self):
        wb = openpyxl.load_workbook(self._tmpl_excel_path)
        sheet = wb.active
        i = 2
        for barcode in self._data:
            sheet.cell(i, 1).value = barcode
            sheet.cell(i, 2).value = self._data.get(barcode)
            i += 1
        wb.save(self._res_excel_path)

    def getPath(self):
        return self._res_excel_path
